from openpyxl import load_workbook
from openpyxl.styles import PatternFill

import os
import win32com.client

def hex_to_vba_color(hex_color):
    hex_color = hex_color.lstrip('#')
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    return r + (g << 8) + (b << 16)

def replace_color(excel_path, old_color, new_color):
    wb = load_workbook(excel_path)

    old_color = old_color.upper().replace("#", "FF")
    new_color = new_color.upper().replace("#", "FF")

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                fill = cell.fill
                if fill and fill.fill_type == 'solid':
                    fg_color = fill.fgColor.rgb
                    # Vérifie s'il y a bien une couleur RGB définie (parfois ce n'est pas le cas)
                    if fg_color and fg_color.upper().startswith(old_color.upper()):
                        cell.fill = PatternFill(start_color=new_color,
                                                end_color=new_color,
                                                fill_type='solid')
    
    wb.save(excel_path)


### With VBA ###
def replace_color2(excel_path, old_color, new_color):
    vba_code = """
Sub ReplaceColors(oldColor As Long, newColor As Long)
    Dim ws As Worksheet
    
    For Each ws In ThisWorkbook.Worksheets
        With Application.FindFormat.Interior
            .PatternColorIndex = xlAutomatic
            .Color = oldColor
            .TintAndShade = 0
            .PatternTintAndShade = 0
        End With
        With Application.ReplaceFormat.Interior
            .PatternColorIndex = xlAutomatic
            .Color = newColor
            .TintAndShade = 0
            .PatternTintAndShade = 0
        End With
        ws.Cells.Replace What:="", Replacement:="", LookAt:=xlPart, SearchOrder:= _
            xlByRows, MatchCase:=False, SearchFormat:=True, ReplaceFormat:=True, _
            FormulaVersion:=xlReplaceFormula2
    Next ws
End Sub
"""
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    long_old_color = hex_to_vba_color(old_color) 
    long_new_color = hex_to_vba_color(new_color) 

    try:
        wb = excel.Workbooks.Open(os.path.abspath(excel_path))
       
        vb_module = wb.VBProject.VBComponents.Add(1)  # 1 = standard module
        vb_module.CodeModule.AddFromString(vba_code)

        # Exécution de la macro
        excel.Application.Run("ReplaceColors", long_old_color, long_new_color)

        wb.Save()
        wb.Close()
    except Exception as e:
        print(f"Erreur : {e}")
    finally:
        excel.Quit()

# replace_color("assets\\test.xlsx",  "#0000FF","#FF00FF")

replace_color2("assets\\test.xlsx", "#FF00FF", "#0000FF")
